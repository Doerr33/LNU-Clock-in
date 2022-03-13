# 实现浏览器打开搜索
# 引入webdriver
import time

import schedule
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

from checkCode import discern_captcha
from checkCode import get_captcha


#数据库版
import pymysql
import pandas as pd

# 判断是否有弹出框
def alert_is_present(driver):
        try:
                alert = driver.switch_to.alert
                alert.text
                return alert
        except:
                return False
def password_is_error(driver):
        try:
                # elepass =  driver.find_element_by_xpath("/html/body/div/div[2]/div[3]/a")
                elepass = driver.find_element(by=By.XPATH, value='/html/body/div/div[2]/div[3]/a')

                return True

        except:
                return False

def job():
        # # 打开excel文件，获取处于活跃状态的表格
        # workbook = load_workbook(filename='num.xlsx')
        # sheet = workbook.active
        # # 获取共多少人
        # print("最大row %d", sheet.max_row)
        # print("最大column %d", sheet.max_column)
        # number = sheet.max_row

        conn = db = pymysql.connect(host="124.222.84.158", port=3306, user="root", password="root", db="book1")
        cur = conn.cursor()

        cur.execute('select title,author,price,school from book1')
        result = cur.fetchall()
        df_result = pd.DataFrame(list(result), columns=["姓名", "账号", "密码","校区"])

        print(result)
        length = len(df_result)
        print(len(df_result))
        print(df_result.head(length))


        for i in range(0,length):
                print(i)
                # 实列化对象，选择启用谷歌，打开浏览器
                driver = webdriver.Chrome()
                driver.implicitly_wait(10)

                # 访问网页
                driver.get("http://tjxx.lnu.edu.cn/login.asp")

                account = df_result.loc[i, '账号']
                password = df_result.loc[i, '密码']
                print("账号" + account)
                print("密码" + password)

                # 输入账号
                # ele = driver.find_element_by_name("userid")
                ele = driver.find_element(by=By.NAME, value='userid')
                ele.send_keys(account)

                # 输入密码
                # ele1 = driver.find_element_by_name("userpwd")
                ele1 = driver.find_element(by=By.NAME, value='userpwd')
                ele1.send_keys(password)


                try:
                        print("尝试try1")
                        get_captcha(driver)
                        captcha = discern_captcha(driver)
                        while(captcha == None or len(captcha) != 4):
                                print("try1 - while获取")
                                driver.find_element(by=By.ID, value='checkimg').click()

                                # driver.find_element_by_id("checkimg").click()
                                get_captcha(driver)
                                captcha = discern_captcha(driver)
                                print("接收结果：", captcha)
                                # 输入验证码

                        # ele9 = driver.find_element_by_name("checkcode")
                        ele9 = driver.find_element(by=By.NAME, value='checkcode')
                        ele9.send_keys(captcha)

                        # driver.find_element_by_id("formSubmitBtn").click()
                        driver.find_element(by=By.ID, value='formSubmitBtn').click()

                        time.sleep(2)
                        # 如果有弹出框 点击确定
                        while alert_is_present(driver):
                                driver.switch_to.alert.accept()
                                time.sleep(2)
                                print("有弹窗")
                                print("alert - while获取")
                                # driver.find_element_by_id("checkimg").click()
                                driver.find_element(by=By.ID, value='checkimg').click()

                                # 输入账号
                                # ele = driver.find_element_by_name("userid")
                                driver.find_element(by=By.NAME, value='userid')
                                ele.send_keys(account)

                                # 输入密码
                                # ele1 = driver.find_element_by_name("userpwd")
                                driver.find_element(by=By.NAME, value='userpwd')
                                ele1.send_keys(password)
                                get_captcha(driver)
                                captcha = discern_captcha(driver)
                                while (captcha == None or len(captcha) != 4):
                                        print("try2 - while获取")
                                        # driver.find_element_by_id("checkimg").click()
                                        driver.find_element(by=By.ID, value='checkimg').click()
                                        get_captcha(driver)
                                        captcha = discern_captcha(driver)
                                        print("接收结果：", captcha)
                                        # 输入验证码

                                # ele9 = driver.find_element_by_name("checkcode")
                                ele9 = driver.find_element(by=By.NAME, value='checkcode')
                                ele9.send_keys(captcha)
                                # driver.find_element_by_id("formSubmitBtn").click()
                                driver.find_element(by=By.ID, value='formSubmitBtn').click()
                                time.sleep(2)
                        #密码错误接受并跳过
                        while password_is_error(driver):
                                driver.switch_to.alert.accept()


                except:
                        driver.quit()
                        print("密码错误")
                        pass
                        continue
                else:
                        time.sleep(1)
                        cellName = df_result.loc[i, '姓名']
                        print(cellName + ":else登录成功")
                        # 选中第一个无

                        # ele4 = driver.find_element_by_xpath("/html/body/div[1]/form/div[3]/div[2]/label[1]/div[1]")
                        ele4 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[3]/div[2]/label[1]/div[1]')
                        driver.execute_script("arguments[0].click();", ele4)

                        # 选中第二个无
                        # ele5 = driver.find_element_by_xpath("/html/body/div[1]/form/div[4]/div[2]/label[1]/div[1]")
                        ele5 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[4]/div[2]/label[1]/div[1]')
                        driver.execute_script("arguments[0].click();", ele5)

                        # 填写温度
                        # loc = driver.find_element_by_id('drtw')  # 定位该元素
                        loc = driver.find_element(by=By.ID, value='drtw')
                        loc.send_keys(Keys.CONTROL + 'a')  # 全选
                        loc.send_keys(Keys.DELETE)  # 删除，清空
                        cell3 = '36.2'
                        # ele2 = driver.find_element_by_name("drtw")
                        ele2 = driver.find_element(by=By.NAME, value='drtw')

                        ele2.send_keys(cell3)

                        # 选中健康
                        # ele6 = driver.find_element_by_xpath("/html/body/div[1]/form/div[6]/div[2]/label[1]/div[1]")
                        ele6 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[6]/div[2]/label[1]/div[1]')
                        driver.execute_script("arguments[0].click();", ele6)

                        # 选中隔离
                        # ele7 = driver.find_element_by_xpath("/html/body/div[1]/form/div[7]/div[2]/label[1]/div[1]")
                        ele7 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[7]/div[2]/label[1]/div[1]')

                        driver.execute_script("arguments[0].click();", ele7)

                        # 是否在校 account = df_result.loc[i, '账号']
                        school = df_result.loc[i,'校区']
                        result = '崇山' in school
                        print(result)
                        if result:
                                # ele8 = driver.find_element_by_xpath("/html/body/div[1]/form/div[8]/div[2]/label[3]/div[1]")
                                ele8 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[8]/div[2]/label[3]/div[1]')

                                driver.execute_script("arguments[0].click();", ele8)
                        else:
                                # ele12 = driver.find_element_by_xpath("/html/body/div[1]/form/div[8]/div[2]/label[2]/div[1]")
                                ele12 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[8]/div[2]/label[2]/div[1]')
                                driver.execute_script("arguments[0].click();", ele12)
                        # 是否无
                        # ele10 = driver.find_element_by_xpath("/html/body/div[1]/form/div[9]/div[2]/label[1]/div[1]")
                        ele10 = driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[9]/div[2]/label[1]/div[1]')
                        driver.execute_script("arguments[0].click();", ele10)

                        time.sleep(30)
                        # # 提交
                        # driver.find_element_by_xpath("/html/body/div[1]/form/div[10]/a").click()
                        driver.find_element(by=By.XPATH, value='/html/body/div[1]/form/div[10]/a').click()

                        # # # 确认提交
                        # driver.find_element_by_xpath("/html/body/div[3]/div[2]/div[2]/a[2]").click()
                        driver.find_element(by=By.XPATH, value='/html/body/div[3]/div[2]/div[2]/a[2]').click()

                        # 退出
                        time.sleep(3)
                        driver.quit()
                        print("提交成功")

# schedule.every(3).seconds.do(job)
schedule.every().day.at("07:11").do(job)
schedule.every().day.at("08:11").do(job)
schedule.every().day.at("09:11").do(job)
schedule.every().day.at("10:11").do(job)

# schedule.every().day.at("08:00").do(job)
while True:
    schedule.run_pending()
time.sleep(3)
